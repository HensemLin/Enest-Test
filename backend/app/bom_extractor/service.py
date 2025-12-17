import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Literal, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from ..config import settings
from ..pdf_processing.service import PDFService
from .models import BomItem
from .schemas import BomItemCreate, BomItemResponse
from .table_parser import BomTableParser


class BomExtractorService:
    """Service for extracting and managing Bill of Materials."""

    def __init__(self, db: Session):
        """
        Initialize BoM extractor service.

        Args:
            db: Database session
        """
        self.db = db
        self.pdf_service = PDFService(db)
        self._ensure_export_dir()

    def _ensure_export_dir(self):
        """Create export directory if it doesn't exist."""
        Path(settings.export_storage_path).mkdir(parents=True, exist_ok=True)

    def extract_bom_items(
        self, pdf_id: int, extraction_mode: str = Literal["auto", "strict", "fuzzy"]
    ) -> tuple[str, List[BomItemResponse]]:
        """
        Extract BoM items from a PDF document.

        Args:
            pdf_id: PDF document ID
            extraction_mode: 'auto', 'strict', or 'fuzzy'

        Returns:
            Tuple of (extraction_job_id, list of BoM items)
        """
        extraction_job_id = str(uuid.uuid4())

        pdf_doc = self.pdf_service.get_pdf_by_id(pdf_id)
        pdf_path = pdf_doc.file_path

        try:
            parser = BomTableParser(extraction_mode=extraction_mode)

            tables = parser.extract_bom_content_from_pdf(pdf_path)

            bom_items = parser.parse_bom_items(tables, pdf_id, extraction_job_id)

            saved_items = []
            for item_data in bom_items:
                bom_item = self._save_bom_item(item_data)
                saved_items.append(BomItemResponse.model_validate(bom_item))

            pdf_doc.status = "ready"
            pdf_doc.last_extraction_date = datetime.now()
            self.db.commit()

            return extraction_job_id, saved_items

        except Exception as e:
            pdf_doc.status = "failed"
            self.db.commit()
            raise e

    def _save_bom_item(self, item_data: dict) -> BomItem:
        """Save a BoM item to the database."""
        item_create = BomItemCreate(**item_data)
        bom_item = BomItem(**item_create.model_dump())

        self.db.add(bom_item)
        self.db.commit()
        self.db.refresh(bom_item)

        return bom_item

    def get_bom_items(
        self,
        pdf_id: Optional[int] = None,
        extraction_job_id: Optional[str] = None,
        skip: int = 0,
        limit: int = 1000,
    ) -> List[BomItem]:
        """
        Get BoM items with optional filters.

        Args:
            pdf_id: Filter by PDF ID
            extraction_job_id: Filter by extraction job
            skip: Number of records to skip
            limit: Maximum number of records

        Returns:
            List of BomItem objects
        """
        query = self.db.query(BomItem)

        if pdf_id is not None:
            query = query.filter(BomItem.pdf_id == pdf_id)

        if extraction_job_id is not None:
            query = query.filter(BomItem.extraction_job_id == extraction_job_id)

        # Order by hierarchy and item number
        query = query.order_by(BomItem.hierarchy_level, BomItem.item_number)

        return query.offset(skip).limit(limit).all()

    def export_to_excel(
        self,
        pdf_id: Optional[int] = None,
        extraction_job_id: Optional[str] = None,
        include_hierarchy: bool = True,
    ) -> tuple[str, str]:
        """
        Export BoM items to Excel with hierarchy formatting.

        Args:
            pdf_id: Optional PDF ID filter
            extraction_job_id: Optional extraction job filter
            include_hierarchy: Whether to format with hierarchy indentation

        Returns:
            Tuple of (file_path, file_name)
        """
        items = self.get_bom_items(pdf_id=pdf_id, extraction_job_id=extraction_job_id)

        if not items:
            raise ValueError("No BoM items found to export")

        data = []
        for item in items:
            row_data = {
                "Item Number": item.item_number or "",
                "Description": item.description,
                "Unit": item.unit or "",
                "Quantity": float(item.quantity) if item.quantity else "",
                "Notes": item.notes or "",
                "Hierarchy Level": item.hierarchy_level,
            }

            if include_hierarchy and item.hierarchy_level > 0:
                indent = "  " * item.hierarchy_level
                row_data["Description"] = f"{indent}{row_data['Description']}"

            data.append(row_data)

        df = pd.DataFrame(data)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if pdf_id:
            file_name = f"bom_pdf_{pdf_id}_{timestamp}.xlsx"
        elif extraction_job_id:
            file_name = f"bom_{extraction_job_id[:8]}_{timestamp}.xlsx"
        else:
            file_name = f"bom_all_{timestamp}.xlsx"

        file_path = os.path.join(settings.export_storage_path, file_name)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Bill of Materials", index=False)

        if include_hierarchy:
            self._apply_hierarchy_formatting(file_path, items)

        return file_path, file_name

    def _apply_hierarchy_formatting(self, file_path: str, items: List[BomItem]):
        """
        Apply hierarchy-based formatting to Excel file.

        Args:
            file_path: Path to Excel file
            items: List of BoM items
        """
        workbook = load_workbook(file_path)
        worksheet = workbook["Bill of Materials"]

        # Define colors for different hierarchy levels
        level_colors = {
            0: "FFFFFF",  # White (top level)
            1: "E8F4F8",  # Light blue
            2: "D0E8F0",  # Medium blue
            3: "B8DCE8",  # Darker blue
        }

        # Apply formatting to each row
        for idx, item in enumerate(items, start=2):  # Start from row 2 (after header)
            level = item.hierarchy_level
            color = level_colors.get(level, "A0D0E0")  # Default color for deep levels

            # Apply background color
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            for cell in worksheet[idx]:
                cell.fill = fill

            # Make top-level items bold
            if level == 0:
                for cell in worksheet[idx]:
                    cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(file_path)

    def delete_bom_items(
        self, pdf_id: Optional[int] = None, extraction_job_id: Optional[str] = None
    ) -> int:
        """
        Delete BoM items with optional filters.

        Args:
            pdf_id: Filter by PDF ID
            extraction_job_id: Filter by extraction job

        Returns:
            Number of items deleted
        """
        query = self.db.query(BomItem)

        if pdf_id is not None:
            query = query.filter(BomItem.pdf_id == pdf_id)

        if extraction_job_id is not None:
            query = query.filter(BomItem.extraction_job_id == extraction_job_id)

        count = query.count()
        query.delete()
        self.db.commit()

        return count
